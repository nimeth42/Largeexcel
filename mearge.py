import openpyxl
import pandas as pd

def process_and_find_similar_data(file1, file2, output_file, chunk_size=100000):
    try:
        wb1 = openpyxl.load_workbook(file1, read_only=True)
        wb2 = openpyxl.load_workbook(file2, read_only=True)
    except FileNotFoundError as e:
        print(f"Error: {e}")
        return

    ws1 = wb1.active
    ws2 = wb2.active

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = 'SimilarData'

    start_row1 = 2
    start_row2 = 2

    while True:
        chunk1 = pd.DataFrame(ws1.iter_rows(min_row=start_row1, max_row=start_row1 + chunk_size, values_only=True))
        chunk2 = pd.DataFrame(ws2.iter_rows(min_row=start_row2, max_row=start_row2 + chunk_size, values_only=True))

        if chunk1.empty and chunk2.empty:
            break

        # Select relevant columns
        chunk1 = chunk1[[0, 9, 10]]  # Adjust indices as needed
        chunk2 = chunk2[[0, 9, 10]]  # Adjust indices as needed

        chunk1.columns = ['CommonColumn', 'DataColumn1', 'DataColumn2']
        chunk2.columns = ['CommonColumn', 'DataColumn1', 'DataColumn2']

        if 'CommonColumn' in chunk1.columns and 'CommonColumn' in chunk2.columns:
            similar_data = pd.merge(chunk1, chunk2, on='CommonColumn')

            for row in similar_data.itertuples(index=False, name=None):
                ws_out.append(row)
        else:
            print("Error: 'CommonColumn' not found in one of the chunks.")
            return

        start_row1 += chunk_size
        start_row2 += chunk_size

    wb_out.save(output_file)

file1 = r'C:\Users\nimeth\Documents\GitHub\Excel read\large_file1.xlsx'
file2 = r'C:\Users\nimeth\Documents\GitHub\Excel read\large_file2.xlsx'
output_file = r'C:\Users\nimeth\Documents\GitHub\Excel read\similar_data.xlsx'

process_and_find_similar_data(file1, file2, output_file)
