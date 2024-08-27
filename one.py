import openpyxl
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows

def process_and_write_large_excel(file1, file2, output_file, chunk_size=100000):
    # Open the Excel files
    try:
        wb1 = openpyxl.load_workbook(file1)
        wb2 = openpyxl.load_workbook(file2)
    except FileNotFoundError as e:
        print(f"Error: {e}")
        return

    ws1 = wb1.active
    ws2 = wb2.active

    # Create a new Excel workbook
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = 'MergedData'

    # Read data in chunks from both Excel files
    def read_chunk(ws, start_row, chunk_size):
        data = []
        for row in ws.iter_rows(min_row=start_row, max_row=start_row + chunk_size, values_only=True):
            data.append(row)
        return pd.DataFrame(data[1:], columns=data[0])

    start_row1 = 2
    start_row2 = 2
    while True:
        chunk1 = read_chunk(ws1, start_row1, chunk_size)
        chunk2 = read_chunk(ws2, start_row2, chunk_size)

        if chunk1.empty and chunk2.empty:
            break

        # Merge the two chunks
        merged_chunk = pd.merge(chunk1, chunk2, on='CommonColumn', how='outer', indicator=True)

        # Filter out rows that are present in both chunks
        unique_data = merged_chunk[merged_chunk['_merge'] != 'both']
        unique_data.drop(columns=['_merge'], inplace=True)

        # Write the unique data to the new workbook
        for row in dataframe_to_rows(unique_data, index=False, header=False):
            ws_out.append(row)

        start_row1 += chunk_size
        start_row2 += chunk_size

    # Save the workbook to a file
    wb_out.save(output_file)

# Example usage with verified paths

